from datetime import datetime
from typing import Optional, List
import csv
from io import StringIO, BytesIO

from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
import motor.motor_asyncio
from bson import ObjectId
import os
from dotenv import load_dotenv
from openai import AsyncOpenAI

load_dotenv()

MONGODB_URI = os.getenv("MONGODB_URI", "mongodb://localhost:27017")
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

app = FastAPI()

if OPENAI_API_KEY:
    ai_client = AsyncOpenAI(api_key=OPENAI_API_KEY)
else:
    ai_client = AsyncOpenAI()

client = motor.motor_asyncio.AsyncIOMotorClient(MONGODB_URI)
db = client.sundeck
prospectos_collection = db.prospectos
etapas_collection = db.etapas
recordatorios_collection = db.recordatorios
escalaciones_collection = db.escalaciones
reagendamientos_collection = db.reagendamientos
comentarios_collection = db.comentarios
training_collection = db.training_materials
progress_collection = db.training_progress
plantillas_collection = db.plantillas

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app.mount("/files", StaticFiles(directory=UPLOAD_DIR), name="files")

class Prospecto(BaseModel):
    nombre: str
    telefono: str
    producto: str
    direccion: Optional[str] = None
    fechaCita: Optional[datetime] = None
    estado: str = "prospecto_nuevo"

class ProspectoDB(Prospecto):
    id: str = Field(default_factory=str, alias="_id")
    createdAt: datetime
    updatedAt: datetime

    class Config:
        allow_population_by_field_name = True
        json_encoders = {ObjectId: str}


class EstadoUpdate(BaseModel):
    estado: str


class Etapa(BaseModel):
    nombre: str
    datos: dict = Field(default_factory=dict)
    archivos: List[str] = Field(default_factory=list)
    fecha: datetime = Field(default_factory=datetime.utcnow)
    usuario: Optional[str] = None


class EtapaDB(Etapa):
    id: str = Field(default_factory=str, alias="_id")
    prospectoId: str

    class Config:
        allow_population_by_field_name = True
        json_encoders = {ObjectId: str}


class EscalacionDB(BaseModel):
    id: str = Field(default_factory=str, alias="_id")
    recordatorioId: str
    prospectoId: str
    cliente: str
    etapa: str
    responsable: str
    fechaVencida: datetime
    prioridad: str
    notificadoA: str
    resuelta: bool = False

    class Config:
        allow_population_by_field_name = True
        json_encoders = {ObjectId: str}


class TrainingItem(BaseModel):
    nombre: str
    categoria: str
    tipo: str
    url: str
    descripcion: Optional[str] = None
    fecha: datetime = Field(default_factory=datetime.utcnow)
    usuario: Optional[str] = None


class TrainingItemDB(TrainingItem):
    id: str = Field(default_factory=str, alias="_id")

    class Config:
        allow_population_by_field_name = True
        json_encoders = {ObjectId: str}


class TrainingProgress(BaseModel):
    trainingId: str
    user: str


class TrainingProgressDB(TrainingProgress):
    id: str = Field(default_factory=str, alias="_id")
    fecha: datetime

    class Config:
        allow_population_by_field_name = True
        json_encoders = {ObjectId: str}


class Plantilla(BaseModel):
    nombre: str
    descripcion: Optional[str] = None
    texto: str
    fechaCreacion: datetime = Field(default_factory=datetime.utcnow)
    creador: Optional[str] = None


class PlantillaDB(Plantilla):
    id: str = Field(default_factory=str, alias="_id")

    class Config:
        allow_population_by_field_name = True
        json_encoders = {ObjectId: str}


class AIMessageRequest(BaseModel):
    nombre: str
    producto: str
    etapa: str
    historial: Optional[str] = None

@app.post("/prospectos", response_model=ProspectoDB)
async def crear_prospecto(prospecto: Prospecto):
    data = prospecto.dict()
    now = datetime.utcnow()
    data.update({"createdAt": now, "updatedAt": now})
    result = await prospectos_collection.insert_one(data)
    data["_id"] = str(result.inserted_id)
    return ProspectoDB(**data)

@app.get("/prospectos", response_model=List[ProspectoDB])
async def listar_prospectos():
    items = []
    async for doc in prospectos_collection.find():
        doc["_id"] = str(doc["_id"])
        items.append(ProspectoDB(**doc))
    return items


@app.put("/prospectos/{id}", response_model=ProspectoDB)
async def actualizar_estado(id: str, data: EstadoUpdate):
    try:
        oid = ObjectId(id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    result = await prospectos_collection.find_one_and_update(
        {"_id": oid},
        {"$set": {"estado": data.estado, "updatedAt": datetime.utcnow()}},
        return_document=True,
    )
    if not result:
        raise HTTPException(status_code=404, detail="Prospecto no encontrado")
    result["_id"] = str(result["_id"])
    return ProspectoDB(**result)


@app.post("/prospectos/{id}/etapas", response_model=EtapaDB)
async def crear_etapa(id: str, etapa: Etapa):
    try:
        oid = ObjectId(id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")

    # verificar que el prospecto exista
    if not await prospectos_collection.find_one({"_id": oid}):
        raise HTTPException(status_code=404, detail="Prospecto no encontrado")

    data = etapa.dict()
    data.update({"prospectoId": oid})
    result = await etapas_collection.insert_one(data)
    data["_id"] = str(result.inserted_id)
    data["prospectoId"] = str(oid)
    return EtapaDB(**data)


@app.get("/prospectos/{id}/etapas", response_model=List[EtapaDB])
async def listar_etapas(id: str):
    try:
        oid = ObjectId(id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")

    items = []
    cursor = etapas_collection.find({"prospectoId": oid})
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        doc["prospectoId"] = str(doc["prospectoId"])
        items.append(EtapaDB(**doc))
    return items


def build_date_filter(field: str, start: Optional[datetime], end: Optional[datetime]):
    filtro = {}
    if start or end:
        rango = {}
        if start:
            rango["$gte"] = start
        if end:
            rango["$lte"] = end
        filtro[field] = rango
    return filtro


@app.get("/metrics/general")
async def metrics_general(
    start: Optional[datetime] = Query(None),
    end: Optional[datetime] = Query(None),
    usuario: Optional[str] = Query(None),
):
    filtro_prospectos = build_date_filter("createdAt", start, end)

    total = await prospectos_collection.count_documents(filtro_prospectos or {})

    pedido_estados = ["pedido", "fabricacion", "instalacion", "entrega", "postventa"]
    pedidos = await prospectos_collection.count_documents(
        {**(filtro_prospectos or {}), "estado": {"$in": pedido_estados}}
    )
    conversion = (pedidos / total * 100) if total else 0

    cotizaciones = await prospectos_collection.count_documents(
        {**(filtro_prospectos or {}), "estado": "cotizacion_activa"}
    )
    instalaciones = await prospectos_collection.count_documents(
        {**(filtro_prospectos or {}), "estado": {"$in": ["instalacion", "entrega", "postventa"]}}
    )

    match_etapas = {"nombre": "Pedido", **build_date_filter("fecha", start, end)}
    if usuario:
        match_etapas["usuario"] = usuario
    cursor = etapas_collection.find(match_etapas)
    diffs = []
    async for e in cursor:
        p = await prospectos_collection.find_one({"_id": e["prospectoId"]})
        if p:
            diffs.append((e["fecha"] - p["createdAt"]).days)
    avg_close = sum(diffs) / len(diffs) if diffs else 0

    return {
        "totalProspectos": total,
        "conversionRate": conversion,
        "avgTimeToClose": avg_close,
        "cotizacionesActivas": cotizaciones,
        "pedidosGenerados": pedidos,
        "instalacionesConfirmadas": instalaciones,
    }


@app.get("/metrics/conversions")
async def metrics_conversions(
    start: Optional[datetime] = Query(None),
    end: Optional[datetime] = Query(None),
):
    filtro = build_date_filter("createdAt", start, end)
    estados = [
        "prospecto_nuevo",
        "cotizacion_activa",
        "pedido",
        "fabricacion",
        "instalacion",
        "entrega",
        "postventa",
    ]
    funnel = []
    for estado in estados:
        count = await prospectos_collection.count_documents({**(filtro or {}), "estado": estado})
        funnel.append({"estado": estado, "count": count})
    conversion_rates = []
    for i in range(len(funnel) - 1):
        current = funnel[i]["count"]
        nxt = funnel[i + 1]["count"]
        rate = (nxt / current * 100) if current else 0
        conversion_rates.append({
            "from": funnel[i]["estado"],
            "to": funnel[i + 1]["estado"],
            "rate": rate,
        })
    return {"funnel": funnel, "conversionRates": conversion_rates}


@app.get("/metrics/user-performance")
async def metrics_user_performance(
    start: Optional[datetime] = Query(None),
    end: Optional[datetime] = Query(None),
    usuario: Optional[str] = Query(None),
):
    match = build_date_filter("fecha", start, end)
    if usuario:
        match["usuario"] = usuario

    total_etapas = await etapas_collection.count_documents(match or {})
    cursor = etapas_collection.aggregate([
        {"$match": match or {}},
        {"$group": {"_id": "$usuario", "count": {"$sum": 1}}},
    ])
    res = []
    async for doc in cursor:
        user = doc["_id"] or "desconocido"
        pct = (doc["count"] / total_etapas * 100) if total_etapas else 0
        res.append({"usuario": user, "porcentaje": pct})
    return res


@app.get("/escalaciones", response_model=List[EscalacionDB])
async def obtener_escalaciones():
    now = datetime.utcnow()
    escalaciones = []
    cursor = recordatorios_collection.find(
        {"resuelto": False, "fechaVencimiento": {"$lt": now}}
    )
    async for rec in cursor:
        dias = (now - rec["fechaVencimiento"]).days
        if dias <= 0:
            continue
        if dias <= 2:
            prioridad = "Normal"
            destino = rec.get("responsable", "")
        elif dias <= 6:
            prioridad = "Urgente"
            destino = "Abigail"
        else:
            prioridad = "Crítico"
            destino = "Admin/CEO"

        esc = await escalaciones_collection.find_one(
            {"recordatorioId": rec["_id"], "resuelta": False}
        )
        prospecto = await prospectos_collection.find_one({"_id": rec["prospectoId"]})
        cliente = prospecto["nombre"] if prospecto else ""
        data = {
            "recordatorioId": rec["_id"],
            "prospectoId": rec["prospectoId"],
            "cliente": cliente,
            "etapa": rec.get("etapa", ""),
            "responsable": rec.get("responsable", ""),
            "fechaVencida": rec["fechaVencimiento"],
            "prioridad": prioridad,
            "notificadoA": destino,
            "resuelta": False,
        }
        if esc:
            # update priority if changed
            await escalaciones_collection.update_one(
                {"_id": esc["_id"]},
                {"$set": {"prioridad": prioridad, "notificadoA": destino}},
            )
            esc.update({"prioridad": prioridad, "notificadoA": destino})
            data = esc
        else:
            result = await escalaciones_collection.insert_one(data)
            data["_id"] = result.inserted_id

        data["_id"] = str(data["_id"])
        data["recordatorioId"] = str(data["recordatorioId"])
        data["prospectoId"] = str(data["prospectoId"])
        escalaciones.append(EscalacionDB(**data))
    return escalaciones


@app.post("/escalaciones/{id}/resolver", response_model=EscalacionDB)
async def resolver_escalacion(id: str):
    try:
        oid = ObjectId(id)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    esc = await escalaciones_collection.find_one_and_update(
        {"_id": oid}, {"$set": {"resuelta": True}}, return_document=True
    )
    if not esc:
        raise HTTPException(status_code=404, detail="Escalación no encontrada")
    esc["_id"] = str(esc["_id"])
    esc["recordatorioId"] = str(esc["recordatorioId"])
    esc["prospectoId"] = str(esc["prospectoId"])
    return EscalacionDB(**esc)


@app.get("/export")
async def exportar(
    start: Optional[datetime] = Query(None),
    end: Optional[datetime] = Query(None),
    estado: Optional[str] = Query(None),
    responsable: Optional[str] = Query(None),
    formato: Optional[str] = Query(None),
):
    filtro = build_date_filter("createdAt", start, end)
    if estado:
        filtro["estado"] = estado
    cursor = prospectos_collection.find(filtro or {})
    registros = []
    async for p in cursor:
        etapa_cursor = (
            await etapas_collection.find({"prospectoId": p["_id"]})
            .sort("fecha", -1)
            .to_list(1)
        )
        etapa_doc = etapa_cursor[0] if etapa_cursor else None
        etapa = etapa_doc["nombre"] if etapa_doc else ""
        resp = etapa_doc.get("usuario", "") if etapa_doc else ""
        if responsable and resp != responsable:
            continue
        escalas = await escalaciones_collection.count_documents(
            {"prospectoId": p["_id"], "resuelta": False}
        )
        registros.append(
            {
                "prospecto": p["nombre"],
                "telefono": p["telefono"],
                "producto": p["producto"],
                "estado": p.get("estado", ""),
                "etapa": etapa,
                "responsable": resp,
                "fechaCita": p.get("fechaCita"),
                "escalaciones": escalas,
                "notas": p.get("direccion", ""),
            }
        )
    if formato == "csv":
        headers = [
            "prospecto",
            "telefono",
            "producto",
            "estado",
            "etapa",
            "responsable",
            "fechaCita",
            "escalaciones",
            "notas",
        ]
        s = StringIO()
        writer = csv.DictWriter(s, fieldnames=headers)
        writer.writeheader()
        writer.writerows(registros)
        s.seek(0)
        return StreamingResponse(
            iter([s.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=export.csv"},
        )
    if formato == "excel":
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no disponible")
        wb = Workbook()
        ws = wb.active
        ws.append(
            [
                "prospecto",
                "telefono",
                "producto",
                "estado",
                "etapa",
                "responsable",
                "fechaCita",
                "escalaciones",
                "notas",
            ]
        )
        for r in registros:
            ws.append(
                [
                    r["prospecto"],
                    r["telefono"],
                    r["producto"],
                    r["estado"],
                    r["etapa"],
                    r["responsable"],
                    r["fechaCita"],
                    r["escalaciones"],
                    r["notas"],
                ]
            )
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=export.xlsx"},
        )
    return registros


@app.get("/reports")
async def reports(
    start: Optional[datetime] = Query(None),
    end: Optional[datetime] = Query(None),
    formato: Optional[str] = Query(None),
):
    filtro = build_date_filter("fecha", start, end)
    reag = []
    async for r in reagendamientos_collection.find(filtro or {}):
        reag.append(
            {
                "cliente": r.get("cliente", ""),
                "producto": r.get("producto", ""),
                "fechaOriginal": r.get("fechaOriginal"),
                "fechaNueva": r.get("fechaNueva"),
                "motivo": r.get("motivo", ""),
                "usuario": r.get("usuario", ""),
            }
        )
    com = []
    async for c in comentarios_collection.find(filtro or {}):
        com.append(
            {
                "cliente": c.get("cliente", ""),
                "producto": c.get("producto", ""),
                "etapa": c.get("etapa", ""),
                "tipo": c.get("tipo", ""),
                "usuario": c.get("usuario", ""),
                "fecha": c.get("fecha"),
            }
        )
    data = {"reagendamientos": reag, "comentarios": com}
    if formato == "csv":
        s = StringIO()
        writer = csv.writer(s)
        writer.writerow(["Reagendamientos"])
        writer.writerow(["cliente", "producto", "fechaOriginal", "fechaNueva", "motivo", "usuario"])
        for r in reag:
            writer.writerow(
                [
                    r["cliente"],
                    r["producto"],
                    r["fechaOriginal"],
                    r["fechaNueva"],
                    r["motivo"],
                    r["usuario"],
                ]
            )
        writer.writerow([])
        writer.writerow(["Comentarios"])
        writer.writerow(["cliente", "producto", "etapa", "tipo", "usuario", "fecha"])
        for c in com:
            writer.writerow(
                [
                    c["cliente"],
                    c["producto"],
                    c["etapa"],
                    c["tipo"],
                    c["usuario"],
                    c["fecha"],
                ]
            )
        s.seek(0)
        return StreamingResponse(
            iter([s.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=reports.csv"},
        )
    if formato == "excel":
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl no disponible")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Reagendamientos"
        ws1.append(["cliente", "producto", "fechaOriginal", "fechaNueva", "motivo", "usuario"])
        for r in reag:
            ws1.append(
                [
                    r["cliente"],
                    r["producto"],
                    r["fechaOriginal"],
                    r["fechaNueva"],
                    r["motivo"],
                    r["usuario"],
                ]
            )
        ws2 = wb.create_sheet("Comentarios")
        ws2.append(["cliente", "producto", "etapa", "tipo", "usuario", "fecha"])
        for c in com:
            ws2.append(
                [
                    c["cliente"],
                    c["producto"],
                    c["etapa"],
                    c["tipo"],
                    c["usuario"],
                    c["fecha"],
                ]
            )
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        return StreamingResponse(
            stream,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=reports.xlsx"},
        )
    return data


@app.post("/training/upload", response_model=TrainingItemDB)
async def upload_training(
    nombre: str = Form(...),
    categoria: str = Form(...),
    usuario: str = Form(""),
    descripcion: Optional[str] = Form(None),
    file: UploadFile = File(...),
):
    filename = f"{int(datetime.utcnow().timestamp())}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    contents = await file.read()
    with open(filepath, "wb") as f:
        f.write(contents)
    data = {
        "nombre": nombre,
        "categoria": categoria,
        "tipo": file.content_type,
        "url": f"/files/{filename}",
        "descripcion": descripcion,
        "fecha": datetime.utcnow(),
        "usuario": usuario or None,
    }
    result = await training_collection.insert_one(data)
    data["_id"] = str(result.inserted_id)
    return TrainingItemDB(**data)


@app.get("/training/list", response_model=List[TrainingItemDB])
async def list_training(
    categoria: Optional[str] = None,
    tipo: Optional[str] = None,
    search: Optional[str] = None,
):
    query = {}
    if categoria:
        query["categoria"] = categoria
    if tipo:
        query["tipo"] = {"$regex": f"^{tipo}"}
    if search:
        query["nombre"] = {"$regex": search, "$options": "i"}
    items: List[TrainingItemDB] = []
    cursor = training_collection.find(query)
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        items.append(TrainingItemDB(**doc))
    return items


@app.get("/training/progress")
async def get_progress(user: str):
    completed: List[str] = []
    cursor = progress_collection.find({"user": user})
    async for doc in cursor:
        completed.append(str(doc["trainingId"]))
    return {"completed": completed}


@app.post("/training/progress", response_model=TrainingProgressDB)
async def set_progress(p: TrainingProgress):
    try:
        oid = ObjectId(p.trainingId)
    except Exception:
        raise HTTPException(status_code=400, detail="ID inválido")
    data = {
        "trainingId": oid,
        "user": p.user,
        "fecha": datetime.utcnow(),
    }
    await progress_collection.update_one(
        {"trainingId": oid, "user": p.user}, {"$set": data}, upsert=True
    )
    doc = await progress_collection.find_one({"trainingId": oid, "user": p.user})
    doc["_id"] = str(doc["_id"])
    doc["trainingId"] = str(doc["trainingId"])
    return TrainingProgressDB(**doc)


@app.post("/plantillas", response_model=PlantillaDB)
async def crear_plantilla(p: Plantilla):
    data = p.dict()
    result = await plantillas_collection.insert_one(data)
    data["_id"] = str(result.inserted_id)
    return PlantillaDB(**data)


@app.get("/plantillas", response_model=List[PlantillaDB])
async def listar_plantillas():
    items: List[PlantillaDB] = []
    cursor = plantillas_collection.find()
    async for doc in cursor:
        doc["_id"] = str(doc["_id"])
        items.append(PlantillaDB(**doc))
    return items


@app.post("/ai/message-suggestion")
async def message_suggestion(req: AIMessageRequest):
    prompt = (
        f"Cliente: {req.nombre}\nProducto: {req.producto}\nEtapa: {req.etapa}. "
    )
    if req.historial:
        prompt += f"Historial: {req.historial}\n"
    prompt += (
        "Genera un mensaje breve, natural y persuasivo en español para contactar al cliente."
    )
    try:
        resp = await ai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "Eres un asistente de ventas de Sundeck.",
                },
                {"role": "user", "content": prompt},
            ],
        )
        mensaje = resp.choices[0].message.content.strip()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"mensaje": mensaje}
